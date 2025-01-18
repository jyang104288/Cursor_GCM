import os
import warnings
import openpyxl
from groq import Groq
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from flask import Flask, request, jsonify
from flask_cors import CORS
from urllib.parse import urlparse

# Suppress warnings
warnings.filterwarnings('ignore')

# Set the API Key as an Environment Variable
os.environ['GROQ_API_KEY'] = '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'
os.environ['USER_AGENT'] = 'GMA_News_AI_Research_Tool/1.0'

# Chatbot Setup
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# File Paths
news_source_list_path = r"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Input\News_Source_List.xlsx"

# List of attributes
attributes = [
    "Title", "Summary", "Affected Products", "Summary Details", "Conformity Assessment",
    "Marking Requirement", "Technical Requirement", "Publish Date", "Effective Date",
    "Mandatory Date", "Consultation Closing"
]

def extract_domain(url):
    parsed_url = urlparse(url)
    return parsed_url.netloc

def validate_url(url):
    domain = extract_domain(url)
    
    # Load the News Source List to check for the domain
    wb = openpyxl.load_workbook(news_source_list_path, data_only=True)
    sheet = wb.active
    
    # Check if the domain is in the first column
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == domain:
            return domain  # Domain is valid
    
    return None  # Domain not found

# Define a function to get the file path dynamically
def get_file_path(domain_name):
    static_path = r"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Input\\"
    # Format the domain name to match the desired file name structure
    formatted_domain_name = f"GMA_News_URL_Input_{domain_name}.xlsx"
    return f"{static_path}{formatted_domain_name}"

def get_text_content(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    return soup.get_text()

def process_url(topic, url, matched_file_name):
    attribute_outputs = {}  # Initialize the dictionary to store prompts and responses

    try:
        # Load the workbook and select the sheets using the dynamic file_path
        file_path = get_file_path(matched_file_name)  # Use the matched_file_name
        print(f"Attempting to load workbook from: {file_path}")  # Debugging line
        wb = openpyxl.load_workbook(file_path, data_only=True)
        output_sheet = wb['Output']

        # Get URL content
        text_content = get_text_content(url)
        if not text_content:
            raise ValueError("Unable to retrieve URL content")
        
        # Process each attribute
        for row_index, output_row in enumerate(output_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not output_row or len(output_row) < 4:
                continue
            
            try:
                attribute, prompt = output_row[0], output_row[1]
                if not attribute or not prompt:
                    continue
                
                # Construct the prompt
                formatted_prompt = (
                    f"As a product regulatory compliance officer, analyze the following text content. "
                    f"Focus on the attribute: '{attribute}'. "
                    f"Specific Question: {prompt} "
                    "Answer the specific question based on the provided text content, but do not repeat the question in the answer. "
                    "Be very concise and give answer directly. When giving a negative answer, simply state 'Not Need'. "
                    "Avoid repeating the question or using multiple sentences to say the same thing. "
                    "Only use information from the provided text content. "
                    "If your answer contains multiple points, please use bullet points and ensure each point is on a new line."
                    "If your answer contains date, please just output the date without description."
                )
                
                # Get AI response
                full_prompt = formatted_prompt + f"\nText content: {text_content[:16000]}... "
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that refines and improves answers."},
                        {"role": "user", "content": full_prompt}
                    ],
                    model="llama3-8b-8192",
                )
                
                final_answer = chat_completion.choices[0].message.content.strip()
                
                # Store the prompt and AI output in the attribute_outputs dictionary
                attribute_outputs[attribute] = {
                    'prompt': formatted_prompt,
                    'ai_output': final_answer
                }
                
            except Exception as row_error:
                attribute_outputs[attribute] = {
                    'error': f"Error processing row {row_index}: {str(row_error)}"
                }
        
    except Exception as e:
        return {"error": str(e)}

    return attribute_outputs

app = Flask(__name__)
CORS(app)

@app.route('/process', methods=['POST'])
def process_request():
    data = request.json
    topic = data.get('topic')
    url = data.get('url')

    if not topic or not url:
        return jsonify({"error": "Topic and URL are required."}), 400

    # Validate the URL and extract the domain
    domain = validate_url(url)
    if not domain:
        return jsonify({"error": "The URL is not validated to work with the AI Extraction."}), 400

    # Correctly determine the input file based on the domain
    matched_file_name = domain  # Use the domain directly, not prefixed
    print(f"Matched file name: {matched_file_name}")  # Debugging line

    try:
        # Call the processing function
        results = process_url(topic, url, matched_file_name)  # Pass topic, url, and matched_file_name
        return jsonify(results), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
