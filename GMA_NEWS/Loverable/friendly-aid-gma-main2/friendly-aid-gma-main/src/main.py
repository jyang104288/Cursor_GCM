from tkinter import ttk
import tkinter as tk
from ttkthemes import ThemedTk
import customtkinter as ctk
from scraping_utils import load_workbook, process_url  # Importing the necessary functions
import ipywidgets as widgets
from IPython.display import display, HTML, clear_output
from datetime import datetime  # Importing datetime module
import openpyxl
import os
import warnings
from groq import Groq
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup

# Suppress warnings
warnings.filterwarnings('ignore')

# Set the API Key as an Environment Variable
os.environ['GROQ_API_KEY'] = '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'

# Chatbot Setup
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# File Paths
input_file_path = r"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Input\GMA_News_URL_Input_CENELEC.xlsx"

# Load the workbook and select the sheets
wb = openpyxl.load_workbook(input_file_path, data_only=True)
output_sheet = wb['Output']

# List of attributes
attributes = [
    "Title", "Summary", "Affected Products", "Summary Details", "Conformity Assessment",
    "Marking Requirement", "Technical Requirement", "Publish Date", "Effective Date",
    "Mandatory Date", "Consultation Closing"
]

def get_text_content(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    return soup.get_text()

def process_url(topic, url, attribute_outputs):
    try:
    
        
        # Load the workbook and select the output sheet
        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        output_sheet = wb['Output']

        # Set output path to the current directory (same as main.py)
        output_path = rf"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Bulk_RAG\Output\{topic}_Output.xlsx"
        
        # Display processing information
        display(HTML(f"""
        <div style='background-color: #f0f0f0; padding: 10px; margin: 10px 0;'>
            <h3>Processing:</h3>
            <p><b>Topic:</b> {topic}</p>
            <p><b>URL:</b> {url}</p>
        </div>
        """))
        
        # Get URL content
        text_content = get_text_content(url)
        if not text_content:
            raise ValueError("Unable to retrieve URL content")
        
        # Process each attribute
        for row_index, output_row in enumerate(output_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not output_row or len(output_row) < 4:  # Ensure there are enough columns
                continue
            
            try:
                # Read attribute and prompt from the correct columns
                attribute = output_row[2]  # Second column (index 2) for attribute
                prompt = output_row[3]      # Third column (index 3) for prompt
                if not attribute or not prompt:
                    continue
                
                # Construct the formatted prompt for the LLM
                formatted_prompt = (
                    f"As a product regulatory compliance officer, analyze the following text content. "
                    f"Focus on the attribute: '{attribute}'. "
                    f"Specific Question: {prompt} "
                    "Answer the specific question based on the provided text content, but do not repeat the question in the answer. "
                    "Be very concise and give answer directly. When giving a negative answer, simply state 'Not Need'. "
                    "Avoid repeating the question or using multiple sentences to say the same thing. "
                    "Only use information from the provided text content. "
                    "If your answer contains multiple points, please use bullet points and ensure each point is on a new line."
                )
                
                # Update the prompt display
                if attribute in attribute_outputs:
                    attribute_outputs[attribute]['prompt'].value = formatted_prompt
                
                # Get AI response
                full_prompt = formatted_prompt + f"\nText content: {text_content[:16000]}... "  # Limit text content to avoid API limits
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that refines and improves answers."},
                        {"role": "user", "content": full_prompt}
                    ],
                    model="llama3-8b-8192",
                )
                
                final_answer = chat_completion.choices[0].message.content.strip()
                
                # Update AI output display
                if attribute in attribute_outputs:
                    attribute_outputs[attribute]['ai_output'].value = final_answer
                
                # Display processing result
                display(HTML(f"""
                <div style='border: 1px solid #ddd; padding: 10px; margin: 10px 0;'>
                    <h4>Attribute: {attribute}</h4>
                    <div style='margin: 10px 0;'>
                        <b>Prompt:</b><br>
                        <pre style='background-color: #f8f8f8; padding: 10px;'>{formatted_prompt}</pre>
                    </div>
                    <div style='margin: 10px 0;'>
                        <b>AI Response:</b><br>
                        <pre style='background-color: #f8f8f8; padding: 10px;'>{final_answer}</pre>
                    </div>
                </div>
                """))
                
            except Exception as row_error:
                display(HTML(f"<div style='color: red;'><b>Error processing row {row_index}:</b> {str(row_error)}</div>"))
        
        # Save results
        wb.save(output_path)
        display(HTML(f"<div style='color: green;'><b>Results saved to:</b> {output_path}</div>"))
        
    except Exception as e:
        display(HTML(f"<div style='color: red;'><b>Error processing URL:</b> {str(e)}</div>"))

class AIResearchTool:
    def __init__(self):
        print("Initializing AIResearchTool...")
        self.root = ThemedTk(theme="arc")  # Modern looking theme
        self.root.title("AI Assisted Research Tool")
        self.root.geometry("1200x800")
        
        # Configure colors and styles
        self.bg_color = "#ffffff"
        self.accent_color = "#646cff"
        self.root.configure(bg=self.bg_color)

        # Initialize log_output as an instance variable
        self.log_output = widgets.Output()
        print("Log output initialized.")

        # Define attributes before using them
        self.attributes = [
            "Title", "Summary", "Affected Products", "Summary Details",
            "Conformity Assessment", "Marking Requirement", "Technical Requirement",
            "Publish Date", "Effective Date", "Mandatory Date", "Consultation Closing"
        ]

        # Initialize attribute_outputs as an empty dictionary
        self.attribute_outputs = {attr: {'prompt': '', 'ai_output': '', 'manual_output': '', 'feedback': '', 'accuracy': 0} for attr in self.attributes}

        # Initialize a variable to hold the output
        self.processed_output = ""

        # Setup UI
        self.setup_ui()

    def setup_ui(self):
        print("Setting up UI...")
        # Header
        header_frame = ttk.Frame(self.root)
        header_frame.pack(pady=20, fill='x')
        
        ttk.Label(
            header_frame, 
            text="AI Assisted Research Tool",
            font=('Helvetica', 24, 'bold')
        ).pack()
        
        ttk.Label(
            header_frame,
            text="Analyze regulatory compliance with advanced AI",
            font=('Helvetica', 12)
        ).pack()

        # Input Section
        input_card = ttk.LabelFrame(self.root, text="Research Parameters")
        input_card.pack(padx=20, pady=10, fill='x')

        # Topic Input
        topic_frame = ttk.Frame(input_card)
        topic_frame.pack(padx=10, pady=5, fill='x')
        ttk.Label(topic_frame, text="Topic").pack(anchor='w')
        self.topic_entry = ttk.Entry(topic_frame)
        self.topic_entry.pack(fill='x')

        # URL Input
        url_frame = ttk.Frame(input_card)
        url_frame.pack(padx=10, pady=5, fill='x')
        ttk.Label(url_frame, text="URL").pack(anchor='w')
        self.url_entry = ttk.Entry(url_frame)
        self.url_entry.pack(fill='x')

        # Buttons
        button_frame = ttk.Frame(input_card)
        button_frame.pack(padx=10, pady=10, fill='x')
        
        clear_btn = ctk.CTkButton(
            button_frame,
            text="Clear",
            command=self.clear_inputs,
            fg_color="transparent",
            border_width=1
        )
        clear_btn.pack(side='right', padx=5)
        
        run_btn = ctk.CTkButton(
            button_frame,
            text="Run",
            command=self.on_button_clicked
        )
        run_btn.pack(side='right', padx=5)

        save_btn = ctk.CTkButton(
            button_frame,
            text="Save",
            command=self.save_changes
        )
        save_btn.pack(side='right', padx=5)

        export_btn = ctk.CTkButton(
            button_frame,
            text="Export",
            command=self.export_to_excel
        )
        export_btn.pack(side='right', padx=5)

        # Analysis Results Section
        results_card = ttk.LabelFrame(self.root, text="Analysis Results")
        results_card.pack(padx=20, pady=10, fill='both', expand=True)

        # Create notebook for tabs
        self.notebook = ttk.Notebook(results_card)
        self.notebook.pack(padx=10, pady=5, fill='both', expand=True)

        # Create tabs for each attribute
        for attr in self.attributes:
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=attr)
            self.create_tab_content(tab)

    def create_tab_content(self, tab):
        print("Creating tab content...")
        # LLM Prompt
        ttk.Label(tab, text="LLM Prompt").pack(anchor='w', padx=10, pady=2)
        prompt_text = tk.Text(tab, height=3)
        prompt_text.pack(fill='x', padx=10, pady=2)

        # AI Generated Analysis
        ttk.Label(tab, text="AI Generated Analysis").pack(anchor='w', padx=10, pady=2)
        ai_analysis = tk.Text(tab, height=4)
        ai_analysis.pack(fill='x', padx=10, pady=2)

        # Expert Analysis
        ttk.Label(tab, text="Expert Analysis").pack(anchor='w', padx=10, pady=2)
        expert_analysis = tk.Text(tab, height=4)
        expert_analysis.pack(fill='x', padx=10, pady=2)

        # Feedback and Accuracy
        feedback_frame = ttk.Frame(tab)
        feedback_frame.pack(fill='x', padx=10, pady=5)

        # Feedback
        feedback_left = ttk.Frame(feedback_frame)
        feedback_left.pack(side='left', fill='x', expand=True, padx=5)
        ttk.Label(feedback_left, text="Feedback").pack(anchor='w')
        feedback_entry = ttk.Entry(feedback_left)
        feedback_entry.pack(fill='x')

        # Accuracy
        feedback_right = ttk.Frame(feedback_frame)
        feedback_right.pack(side='left', fill='x', expand=True, padx=5)
        ttk.Label(feedback_right, text="Accuracy (%)").pack(anchor='w')
        accuracy_entry = ttk.Entry(feedback_right)
        accuracy_entry.pack(fill='x')

        # Store references to widgets
        tab_name = self.notebook.tab(tab)['text']
        self.attribute_outputs[tab_name].update({
            'prompt_widget': prompt_text,
            'ai_output_widget': ai_analysis,
            'manual_output_widget': expert_analysis,
            'feedback_widget': feedback_entry,
            'accuracy_widget': accuracy_entry
        })

    def clear_inputs(self):
        print("Clearing inputs...")
        self.topic_entry.delete(0, tk.END)
        self.url_entry.delete(0, tk.END)
        
        display(HTML("<b>Inputs cleared.</b>"))
        with self.log_output:
            clear_output()
        # Clear attribute outputs
        for attr in self.attributes:
            attribute_outputs[attr]['prompt'].value = ''
            attribute_outputs[attr]['ai_output'].value = ''
            attribute_outputs[attr]['manual_output'].value = ''
            attribute_outputs[attr]['feedback'].value = ''
            attribute_outputs[attr]['accuracy'].value = 0

    def on_button_clicked(self):
        print("Run button clicked.")
        with self.log_output:
            clear_output()  # Clear previous output
            
            try:
                # Get input values
                topic = self.topic_entry.get()
                url = self.url_entry.get()
                print(f"Topic: {topic}, URL: {url}")
                
                # Input validation
                if not topic:
                    display(HTML("<div style='color: red;'><b>Error:</b> Please enter a topic</div>"))
                    return
                if not url:
                    display(HTML("<div style='color: red;'><b>Error:</b> Please enter a URL</div>"))
                    return
                    
                # Display processing information
                display(HTML("<div style='color: blue;'><b>Processing...</b></div>"))
                
                # Process URL and update attribute outputs
                updated_outputs = process_url(topic, url, self.attribute_outputs)
                self.attribute_outputs.update(updated_outputs)
                
                # Update the UI with results
                for attr, outputs in self.attribute_outputs.items():
                    if 'prompt_widget' in outputs:
                        # Update prompt
                        outputs['prompt_widget'].delete(1.0, tk.END)
                        outputs['prompt_widget'].insert(tk.END, outputs.get('prompt', ''))
                        
                        # Update AI output
                        outputs['ai_output_widget'].delete(1.0, tk.END)
                        outputs['ai_output_widget'].insert(tk.END, outputs.get('ai_output', ''))
                        
                        print(f"Updated tab {attr} with prompt and AI output")
                
                # Display success message
                display(HTML("<div style='color: green;'><b>Processing completed successfully!</b></div>"))
                
            except Exception as e:
                print(f"Error during processing: {str(e)}")
                display(HTML(f"<div style='color: red;'><b>Error during processing:</b> {str(e)}</div>"))

    def save_changes(self):
        print("Saving changes...")
        with self.log_output:
            clear_output()
        for attr, outputs in attribute_outputs.items():
            row_index = self.attributes.index(attr) + 2
            output_sheet.cell(row=row_index, column=response_col_index, value=outputs['prompt'].value)
            output_sheet.cell(row=row_index, column=response_col_index, value=outputs['ai_output'].value)
            output_sheet.cell(row=row_index, column=response_col_index, value=outputs['manual_output'].value)
            output_sheet.cell(row=row_index, column=response_col_index + 1, value=outputs['feedback'].value)
            output_sheet.cell(row=row_index, column=response_col_index + 2, value=outputs['accuracy'].value)
        wb.save(file_path)
        display(HTML("<b>Changes saved successfully!</b>"))

    def export_to_excel(self):
        print("Exporting to Excel...")
        with self.log_output:
            clear_output()
        export_data = []
        for attr, outputs in attribute_outputs.items():
            export_data.append([
                attr,
                outputs['prompt'].value,
                outputs['ai_output'].value,
                outputs['manual_output'].value,
                outputs['feedback'].value,
                outputs['accuracy'].value
            ])
        df = pd.DataFrame(export_data, columns=['Attribute', 'Content', 'AI Output', 'Manual Output', 'Feedback', 'Accuracy (%)'])
        wb_export = Workbook()
        ws = wb_export.active
        ws.title = "Exported"

        # Write headers
        headers = ['Attribute', 'LLM Prompt', 'AI Generated', 'Expert Created', 'Feedback', 'Accuracy (%)']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Write data
        for r, row in enumerate(df.values, start=2):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

        # Create a root window and hide it
        root = tk.Tk()
        root.withdraw()

        # Ask user for export location and filename
        export_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel File"
        )

        if not export_path:
            display(HTML("<b>Export cancelled.</b>"))
            return

        # Save the workbook
        wb_export.save(export_path)
        display(HTML(f"<b>Data exported successfully to:</b> {export_path}"))

    def run(self):
        print("Running the application...")
        self.root.mainloop()

if __name__ == "__main__":
    app = AIResearchTool()
    app.run()