import os
import openpyxl
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from groq import Groq
from IPython.display import display, HTML

# Set the API Key as an Environment Variable
os.environ['GROQ_API_KEY'] = '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'

# Chatbot Setup
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

def load_workbook(file_path):
    """Load the Excel workbook and return the output sheet."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    return wb['Output']

def get_text_content(url):
    """Fetch and return the text content from the given URL."""
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    return soup.get_text()

def process_url(topic, url, attribute_outputs):
    print(f"Starting process_url with topic: {topic}, url: {url}")
    try:
        # Set output path to the current directory (same as main.py)
        output_dir = os.path.dirname(os.path.abspath(__file__))  # Get the directory of the current file
        output_path = os.path.join(output_dir, f"{topic}_Output.xlsx")  # Updated to remove date_str

        llm_prompt_column = 'LLM Prompt'
        response_column = 'AI Generated Answer'
        header_row = output_sheet[1]
        response_col_index = next((i for i, cell in enumerate(header_row, 1) if cell.value == response_column), None)
        llm_prompt_col_index = next((i for i, cell in enumerate(header_row, 1) if cell.value == llm_prompt_column), None)
        
        if not response_col_index or not llm_prompt_col_index:
            raise ValueError(f"Required columns {llm_prompt_column} and {response_column} not found in Output sheet.")
        
        # Extract text content from the URL
        text_content = get_text_content(url)
        print(f"Fetched text content from URL: {len(text_content)} characters long.")
        if not text_content:
            raise ValueError("Unable to fetch URL content")
            
        # Process each attribute
        for row_index, output_row in enumerate(output_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not output_row or len(output_row) < 4:
                continue
                
            try:
                attribute, prompt = output_row[0], output_row[1]
                if not attribute or not prompt:
                    continue
                    
                # Construct the prompt
                formatted_prompt = (
                    f"As a product regulatory compliance officer, analyze the following text content about {topic}. "
                    f"Focus on the attribute: '{attribute}'. "
                    f"Specific Question: {prompt} "
                    "Answer the specific question based on the provided text content, but do not repeat the question in the answer. "
                    "Be very concise and give answer directly. When giving a negative answer, simply state 'Not Need'. "
                    "Avoid repeating the question or using multiple sentences to say the same thing. "
                    "Only use information from the provided text content. "
                    "If your answer contains multiple points, please use bullet points and ensure each point is on a new line."
                    "If your answer contains date, please just output the date without description."
                )
                
                print(f"Sending LLM request for attribute: {attribute}")
                print(f"Formatted prompt: {formatted_prompt[:100]}...")  # Log the first 100 characters of the prompt
                
                # Update prompt display
                if attribute in attribute_outputs:
                    attribute_outputs[attribute]['prompt'].value = formatted_prompt
                
                # Get AI response
                full_prompt = formatted_prompt + f"\nText content: {text_content[:20000]}... "
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that refines and improves answers."},
                        {"role": "user", "content": full_prompt}
                    ],
                    model="llama3-8b-8192",
                )
                
                final_answer = chat_completion.choices[0].message.content.strip()
                print(f"Received response for attribute: {attribute}, response length: {len(final_answer)} characters.")
                
                # Update AI output display
                if attribute in attribute_outputs:
                    attribute_outputs[attribute]['ai_output'].value = final_answer
                
                # Display processing result
                display(HTML(f"<div style='border: 1px solid #ddd; padding: 10px; margin: 10px 0;'><h4>Attribute: {attribute}</h4><div style='margin: 10px 0;'><b>Prompt:</b><br><pre style='background-color: #f8f8f8; padding: 10px;'>{formatted_prompt}</pre></div><div style='margin: 10px 0;'><b>AI Response:</b><br><pre style='background-color: #f8f8f8; padding: 10px;'>{final_answer}</pre></div></div>"))
                
            except Exception as row_error:
                print(f"Error processing row {row_index}: {str(row_error)}")
                display(HTML(f"<div style='color: red;'><b>Error processing row {row_index}:</b> {str(row_error)}</div>"))
        
        # Save results
        wb.save(output_path)
        print(f"Results saved to: {output_path}")
        display(HTML(f"<div style='color: green;'><b>Results saved to:</b> {output_path}</div>"))
        
    except Exception as e:
        print(f"Error processing URL: {str(e)}")
        display(HTML(f"<div style='color: red;'><b>Error processing URL:</b> {str(e)}</div>"))