import os
import warnings
import openpyxl
from groq import Groq
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import messagebox, scrolledtext
import pandas as pd

# Suppress warnings
warnings.filterwarnings('ignore')

# Set the API Key as an Environment Variable
os.environ['GROQ_API_KEY'] = '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'

# Set USER_AGENT environment variable
os.environ['USER_AGENT'] = 'GMA_News_AI_Research_Tool/1.0'

# Chatbot Setup
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# File Paths
file_path = r"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Input\GMA_News_URL_Input_CENELEC.xlsx"

# Load the workbook and select the sheets
wb = openpyxl.load_workbook(file_path, data_only=True)
output_sheet = wb['Output']

def get_text_content(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    return soup.get_text()

def process_url(topic, url):
    try:
        output_path = rf"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Bulk_RAG\Output\{topic}_Output.xlsx"
        
        text_content = get_text_content(url)
        if not text_content:
            raise ValueError("Unable to retrieve URL content")
        
        for row_index, output_row in enumerate(output_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not output_row or len(output_row) < 4:
                continue
            
            try:
                attribute, prompt = output_row[0], output_row[1]
                if not attribute or not prompt:
                    continue
                
                formatted_prompt = (
                    f"As a product regulatory compliance officer, analyze the following text content. "
                    f"Focus on the attribute: '{attribute}'. "
                    f"Specific Question: {prompt} "
                    "Answer the specific question based on the provided text content, but do not repeat the question in the answer. "
                    "Be very concise and give answer directly. When giving a negative answer, simply state 'Not Need'. "
                    "Avoid repeating the question or using multiple sentences to say the same thing. "
                    "Only use information from the provided text content. "
                    "If your answer contains multiple points, please use bullet points and ensure each point is on a new line."
                    "If your answer contains date, please just output the date without description."
                )
                
                # Get AI response
                full_prompt = formatted_prompt + f"\nText content: {text_content[:16000]}... "
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that refines and improves answers."},
                        {"role": "user", "content": full_prompt}
                    ],
                    model="llama3-8b-8192",
                )
                
                final_answer = chat_completion.choices[0].message.content.strip()
                
                # Save results to output sheet
                output_sheet.cell(row=row_index, column=3, value=final_answer)  # Assuming column 3 is for AI output
                
            except Exception as row_error:
                print(f"Error processing row {row_index}: {str(row_error)}")
        
        # Save results
        wb.save(output_path)
        messagebox.showinfo("Success", f"Results saved to: {output_path}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Error processing URL: {str(e)}")

def on_submit():
    topic = topic_entry.get()
    url = url_entry.get()
    if topic and url:
        process_url(topic, url)
    else:
        messagebox.showwarning("Input Error", "Please enter both topic and URL.")

# Create the main window
root = tk.Tk()
root.title("AI Assisted Research Tool")
root.geometry("800x600")  # Increased size for better visibility
root.configure(bg="#f9fafb")  # Set a light background color

# Research Parameters Frame
params_frame = tk.Frame(root, bg="#f9fafb")
params_frame.pack(pady=20)

tk.Label(params_frame, text="Research Parameters", bg="#f9fafb", font=("Arial", 16)).grid(row=0, columnspan=2, pady=10)

tk.Label(params_frame, text="Topic:", bg="#f9fafb", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=5)
topic_entry = tk.Entry(params_frame, width=50, font=("Arial", 12))
topic_entry.grid(row=1, column=1, padx=10, pady=5)

tk.Label(params_frame, text="URL:", bg="#f9fafb", font=("Arial", 12)).grid(row=2, column=0, padx=10, pady=5)
url_entry = tk.Entry(params_frame, width=50, font=("Arial", 12))
url_entry.grid(row=2, column=1, padx=10, pady=5)

# Buttons
button_frame = tk.Frame(root, bg="#f9fafb")
button_frame.pack(pady=10)

clear_button = tk.Button(button_frame, text="Clear", command=lambda: [topic_entry.delete(0, tk.END), url_entry.delete(0, tk.END)], bg="#e2e8f0", font=("Arial", 12))
clear_button.grid(row=0, column=0, padx=10)

analyze_button = tk.Button(button_frame, text="Analyze", command=on_submit, bg="#4f46e5", fg="white", font=("Arial", 12))
analyze_button.grid(row=0, column=1, padx=10)

# Analysis Results Frame
results_frame = tk.Frame(root, bg="#f9fafb")
results_frame.pack(pady=20, fill=tk.BOTH, expand=True)

tk.Label(results_frame, text="Analysis Results", bg="#f9fafb", font=("Arial", 16)).grid(row=0, columnspan=2, pady=10)

# Create a canvas for scrolling
canvas = tk.Canvas(results_frame, bg="#f9fafb")
scrollbar = tk.Scrollbar(results_frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas, bg="#f9fafb")

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

# Configure scrollbar
canvas.configure(yscrollcommand=scrollbar.set)

# Pack the canvas and scrollbar
canvas.grid(row=1, column=0, sticky="nsew")
scrollbar.grid(row=1, column=1, sticky="ns")

# Create scrolled text areas for results
tk.Label(scrollable_frame, text="Title Analysis", bg="#f9fafb", font=("Arial", 12)).grid(row=0, columnspan=2, pady=5)

tk.Label(scrollable_frame, text="LLM Prompt:", bg="#f9fafb").grid(row=1, column=0, sticky='w', padx=10)
llm_prompt_text = scrolledtext.ScrolledText(scrollable_frame, width=60, height=5, font=("Arial", 10))
llm_prompt_text.grid(row=1, column=1, padx=10)

tk.Label(scrollable_frame, text="AI Generated Analysis:", bg="#f9fafb").grid(row=2, column=0, sticky='w', padx=10)
ai_analysis_text = scrolledtext.ScrolledText(scrollable_frame, width=60, height=5, font=("Arial", 10))
ai_analysis_text.grid(row=2, column=1, padx=10)

tk.Label(scrollable_frame, text="Expert Analysis:", bg="#f9fafb").grid(row=3, column=0, sticky='w', padx=10)
expert_analysis_text = scrolledtext.ScrolledText(scrollable_frame, width=60, height=5, font=("Arial", 10))
expert_analysis_text.grid(row=3, column=1, padx=10)

tk.Label(scrollable_frame, text="Feedback:", bg="#f9fafb").grid(row=4, column=0, sticky='w', padx=10)
feedback_text = scrolledtext.ScrolledText(scrollable_frame, width=60, height=3, font=("Arial", 10))
feedback_text.grid(row=4, column=1, padx=10)

tk.Label(scrollable_frame, text="Accuracy (%):", bg="#f9fafb").grid(row=5, column=0, sticky='w', padx=10)
accuracy_entry = tk.Entry(scrollable_frame, width=20, font=("Arial", 10))
accuracy_entry.grid(row=5, column=1, padx=10)

# Start the GUI event loop
root.mainloop()

def read_input_file(file_path):
    # Assuming the input file is a CSV
    data = pd.read_csv(file_path)
    return data

def form_prompt(data):
    # Logic to form the prompt based on the data
    prompts = []
    for index, row in data.iterrows():
        # Example: Create a prompt based on a specific column
        prompt = f"Analyze the following: {row['attribute_column']}"  # Adjust as needed
        prompts.append(prompt)
    return prompts

def process_url(url):
    # Logic to process the URL and retrieve information
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        # Extract relevant information from the soup object
        # Example: title = soup.title.string
        return soup.title.string  # Adjust based on what you need
    else:
        return None