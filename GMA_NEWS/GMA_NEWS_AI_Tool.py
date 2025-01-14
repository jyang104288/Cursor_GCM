import os
import warnings
import openpyxl
from groq import Groq
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import messagebox, scrolledtext
from tkinter import ttk

# Suppress warnings
warnings.filterwarnings('ignore')

# Set the API Key as an Environment Variable
os.environ['GROQ_API_KEY'] = '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'

# Set USER_AGENT environment variable
os.environ['USER_AGENT'] = 'GMA_News_AI_Research_Tool/1.0'

# Chatbot Setup
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# File Paths
file_path = r"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Input\GMA_News_URL_Input_CENELEC.xlsx"

# Load the workbook and select the sheets
wb = openpyxl.load_workbook(file_path, data_only=True)
output_sheet = wb['Output']

# List of attributes
attributes = [
    "Title", "Summary", "Affected Products", "Summary Details", "Conformity Assessment",
    "Marking Requirement", "Technical Requirement", "Publish Date", "Effective Date",
    "Mandatory Date", "Consultation Closing"
]

def get_text_content(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    return soup.get_text()

def process_url(topic, url):
    try:
        output_path = rf"C:\Users\104288\UL Solutions\GMA - Global Market Access - AI POC\GMA News\Bulk_RAG\Output\{topic}_Output.xlsx"
        
        text_content = get_text_content(url)
        if not text_content:
            raise ValueError("Unable to retrieve URL content")
        
        for row_index, output_row in enumerate(output_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not output_row or len(output_row) < 4:
                continue
            
            try:
                attribute, prompt = output_row[0], output_row[1]
                if not attribute or not prompt:
                    continue
                
                formatted_prompt = (
                    f"As a product regulatory compliance officer, analyze the following text content. "
                    f"Focus on the attribute: '{attribute}'. "
                    f"Specific Question: {prompt} "
                    "Answer the specific question based on the provided text content, but do not repeat the question in the answer. "
                    "Be very concise and give answer directly. When giving a negative answer, simply state 'Not Need'. "
                    "Avoid repeating the question or using multiple sentences to say the same thing. "
                    "Only use information from the provided text content. "
                    "If your answer contains multiple points, please use bullet points and ensure each point is on a new line."
                    "If your answer contains date, please just output the date without description."
                )
                
                # Get AI response
                full_prompt = formatted_prompt + f"\nText content: {text_content[:16000]}... "
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that refines and improves answers."},
                        {"role": "user", "content": full_prompt}
                    ],
                    model="llama3-8b-8192",
                )
                
                final_answer = chat_completion.choices[0].message.content.strip()
                
                # Save results to output sheet
                output_sheet.cell(row=row_index, column=3, value=final_answer)  # Assuming column 3 is for AI output
                
            except Exception as row_error:
                print(f"Error processing row {row_index}: {str(row_error)}")
        
        # Save results
        wb.save(output_path)
        messagebox.showinfo("Success", f"Results saved to: {output_path}")
        
    except Exception as e:
        messagebox.showerror("Error", f"Error processing URL: {str(e)}")

def on_submit():
    topic = topic_entry.get()
    url = url_entry.get()
    if topic and url:
        process_url(topic, url)
    else:
        messagebox.showwarning("Input Error", "Please enter both topic and URL.")

def clear_fields():
    topic_entry.delete(0, tk.END)
    url_entry.delete(0, tk.END)
    for tab in tabs.winfo_children():
        llm_prompt_text = tab.children['llm_prompt_text']
        ai_output_text = tab.children['ai_output_text']
        expert_output_text = tab.children['expert_output_text']
        feedback_text = tab.children['feedback_text']
        accuracy_text = tab.children['accuracy_text']
        llm_prompt_text.delete(1.0, tk.END)
        ai_output_text.delete(1.0, tk.END)
        expert_output_text.delete(1.0, tk.END)
        feedback_text.delete(0, tk.END)
        accuracy_text.delete(0, tk.END)

# Create the main window
root = tk.Tk()
root.title("AI Assisted Research Tool")
root.geometry("1000x800")  # Set a larger initial window size
root.configure(bg="#f0f0f0")

# Allow resizing
root.resizable(True, True)  # Enable resizing

# Load the Azure theme
root.tk.call("source", "azure.tcl")  # Ensure azure.tcl is in the same directory
root.tk.call("set_theme", "light")  # Set the initial theme

# Header
header_frame = ttk.Frame(root, padding="10")
header_frame.pack(fill="x")

title_label = ttk.Label(header_frame, text="AI Assisted Research Tool", font=("Arial", 24, "bold"))
title_label.pack(pady=5)

subtitle_label = ttk.Label(header_frame, text="Analyze regulatory compliance with advanced AI", font=("Arial", 14))
subtitle_label.pack(pady=5)

# Research Parameters Section
param_frame = ttk.LabelFrame(root, text="Research Parameters", padding="10")
param_frame.pack(pady=10, fill="x", padx=20)  # Added padding on the x-axis

ttk.Label(param_frame, text="Topic:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
topic_entry = ttk.Entry(param_frame, width=80)
topic_entry.grid(row=0, column=1, padx=5, pady=5)

ttk.Label(param_frame, text="URL:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
url_entry = ttk.Entry(param_frame, width=80)
url_entry.grid(row=1, column=1, padx=5, pady=5)

button_frame = ttk.Frame(param_frame)
button_frame.grid(row=2, columnspan=2, pady=10)

run_button = ttk.Button(button_frame, text="Analyze", command=on_submit)
run_button.pack(side=tk.LEFT, padx=5)

clear_button = ttk.Button(button_frame, text="Clear", command=clear_fields)
clear_button.pack(side=tk.LEFT, padx=5)

# Analysis Results Section
results_frame = ttk.LabelFrame(root, text="Analysis Results", padding="10")
results_frame.pack(pady=10, fill="both", expand=True, padx=20)  # Added padding on the x-axis

# Create tabs for each attribute
tabs = ttk.Notebook(results_frame)
tabs.pack(pady=10, expand=True)

for attribute in attributes:
    tab = ttk.Frame(tabs)
    tabs.add(tab, text=attribute)

    # Create a frame for the text area and scrollbar
    text_frame = ttk.Frame(tab)
    text_frame.pack(pady=5, fill="both", expand=True)

    # Create a scrollbar for the entire tab
    scrollbar = ttk.Scrollbar(text_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Create a canvas to hold the text areas
    canvas = tk.Canvas(text_frame)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Create a frame inside the canvas
    inner_frame = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=inner_frame, anchor='nw')

    # Link scrollbar to canvas
    def on_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    inner_frame.bind("<Configure>", on_configure)

    # LLM Prompt
    llm_prompt_label = ttk.Label(inner_frame, text="LLM Prompt:")
    llm_prompt_label.pack(pady=5)
    llm_prompt_text = scrolledtext.ScrolledText(inner_frame, width=80, height=5, borderwidth=2, relief="groove")
    llm_prompt_text.pack(pady=5, fill="both", expand=True)

    # AI Generated News Attributes
    ai_output_label = ttk.Label(inner_frame, text="AI Generated Analysis:")
    ai_output_label.pack(pady=5)
    ai_output_text = scrolledtext.ScrolledText(inner_frame, width=80, height=5, borderwidth=2, relief="groove")
    ai_output_text.pack(pady=5, fill="both", expand=True)

    # Expert Analysis
    expert_output_label = ttk.Label(inner_frame, text="Expert Analysis:")
    expert_output_label.pack(pady=5)
    expert_output_text = scrolledtext.ScrolledText(inner_frame, width=80, height=5, borderwidth=2, relief="groove")
    expert_output_text.pack(pady=5, fill="both", expand=True)

    # Feedback
    feedback_label = ttk.Label(inner_frame, text="Feedback:")
    feedback_label.pack(pady=5)
    feedback_text = ttk.Entry(inner_frame, width=80)
    feedback_text.pack(pady=5)

    # Accuracy
    accuracy_label = ttk.Label(inner_frame, text="Accuracy (%):")
    accuracy_label.pack(pady=5)
    accuracy_text = ttk.Entry(inner_frame, width=80)
    accuracy_text.pack(pady=5)

    # Configure scrollbar to control the canvas
    scrollbar.config(command=canvas.yview)
    canvas.config(yscrollcommand=scrollbar.set)

# Start the GUI event loop
root.mainloop()